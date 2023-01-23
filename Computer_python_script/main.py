import json
import socket
from openpyxl import load_workbook, Workbook
import os

IP = 
PORT = 


# dealing with excel stuff
def excel_stuff(json_file):
    # reading the important information from a json file
    item_id = json_file.get("id")
    item_Name = json_file.get("Item Name")
    item_location = json_file.get("Location")

    # name of the database file
    database_name = 'DataBase.xlsx'

    # check if the database file exists
    # if database exists
    if os.path.exists(database_name):
        # open the database file
        workbook = load_workbook(database_name)

        # Get the active sheet
        sheet = workbook.active

        exists = False
        # checking the database to see if the item being entered exists
        # range is starts at 2 to skip the Title Row and includes the last row (why use max+1)
        for row in range(2, sheet.max_row+1):
            if sheet['A'+str(row)].value == item_id:
                sheet['B'+str(row)] = item_Name
                sheet['C' + str(row)] = item_location
                exists = True
                print(f"updated Item on line {row}")
                break
        # check if the item being entered exists
        if not exists:
            New_row = str(sheet.max_row+1)
            sheet['A' + New_row] = item_id
            sheet['B' + New_row] = item_Name
            sheet['C' + New_row] = item_location
            print(f"Created Item on line {New_row}")

        # Save the changes
        workbook.save(database_name)

    # if database doesnt exist
    else:
        print("created database")

        # Create a new Excel workbook
        workbook = Workbook()

        # Get the active sheet
        sheet = workbook.active

        # Adding the top line of the database
        sheet['A1'] = 'ID'
        sheet['B1'] = 'Item Name'
        sheet['C1'] = 'Location'

        # adding the first item from the json file
        sheet['A2'] = item_id
        sheet['B2'] = item_Name
        sheet['C2'] = item_location

        # Save the workbook to a file
        workbook.save(database_name)


# dealing with json stuff
def receive_json():
    # Create a socket
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

    # Bind the socket to an address and port
    sock.bind((IP, PORT))

    # Listen for incoming connections
    sock.listen(1)

    while True:
        # Accept a connection
        connection, address = sock.accept()

        # Receive data
        data = b''
        while True:
            part = connection.recv(1024)
            data += part
            if len(part) < 1024:
                break
        # decode the bytes to string
        json_data = json.loads(data.decode())
        print(json_data)

        excel_stuff(json_data)

        # Close the connection
        connection.close()

if __name__ == '__main__':
    receive_json()
